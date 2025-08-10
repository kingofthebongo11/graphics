import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import os
from pathlib import Path
import ast

import numpy as np

import logging

from mylibproject.Figurenameclass import FigureNames

from mylibproject.myutils import to_percent

from mylibproject.myutils_work_with_file import extract_data_frec_from_file

from mylibproject.myutils_widgets import make_context_menu, select_path, message_log, add_hotkeys
from tabs.tab3 import create_tab3

from tabs.tab2 import create_tab2

from tabs.tab1 import create_tab1


logger = logging.getLogger(__name__)





def configure_matplotlib():
    """Configure global matplotlib settings used across the application."""
    plt.rcParams['font.family'] = 'Times New Roman'
    plt.rcParams['font.size'] = 14  # Общий размер шрифта
    plt.rcParams['axes.titlesize'] = 14  # Размер шрифта для заголовков осей
    plt.rcParams['axes.labelsize'] = 14  # Размер шрифта для подписей осей
    plt.rcParams['xtick.labelsize'] = 14  # Размер шрифта для меток оси X
    plt.rcParams['ytick.labelsize'] = 14  # Размер шрифта для меток оси Y
    plt.rcParams['legend.fontsize'] = 14  # Размер шрифта для легенды





def create_text(parent, method='text', height=10, wrap='word', state='normal', scrollbar=False, max_lines=0):
    """Создает текстовое поле с возможностью прокрутки и контекстного меню."""
    # Создаем текстовое поле с заданными параметрами
    if method == 'entry':
        text_widget = tk.Entry(parent, state=state)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Заполнение области фрейма

    elif method == 'text':
        text_widget = tk.Text(parent, height=height, wrap=wrap, state=state)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Заполнение области фрейма
        if max_lines > 0:
            def limit_text_lines(text_widget, max_lines):
                """Ограничивает количество строк в текстовом поле."""

                def check_lines(event):
                    if max_lines > 0:
                        current_text = text_widget.get("1.0", tk.END)  # Получаем текущий текст
                        lines = current_text.splitlines()  # Разбиваем на строки
                        if len(lines) > max_lines:  # Если строк больше чем допустимо
                            text_widget.delete(f"{max_lines + 1}.0", tk.END)  # Удаляем лишние строки

                text_widget.bind("<KeyRelease>", check_lines)  # Проверяем после каждой клавиши
            limit_text_lines(text_widget, max_lines)  # Применяем ограничение по количеству строк

        if scrollbar:
            def bind_scrollbar(log_text, log_scrollbar):
                def update_scrollbar(log_text, log_scrollbar):
                    # Обновление позиции и размера скроллбара в зависимости от содержимого текстового поля
                    log_scrollbar.config(command=log_text.yview)
                    log_text['yscrollcommand'] = log_scrollbar.set

                # Привязываем события к текстовому полю и скроллбару
                log_text.bind("<KeyRelease>", lambda event: update_scrollbar(log_text, log_scrollbar))
                log_text.bind("<MouseWheel>", lambda event: update_scrollbar(log_text, log_scrollbar))
                log_text.bind("<Configure>", lambda event: update_scrollbar(log_text, log_scrollbar))
            # Создаем и размещаем скроллбар, если требуется
            text_scrollbar = ttk.Scrollbar(parent, command=text_widget.yview)
            text_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  # Размещение скроллбара справа
            text_widget['yscrollcommand'] = text_scrollbar.set  # Привязываем скроллбар к текстовому полю
            bind_scrollbar(text_widget, text_scrollbar)  # Привязываем события для скроллбара
    else:
        raise ValueError("Некорректное значение параметра method. Должно быть 'entry' или 'text'.")
    # Применяем контекстное меню к текстовому полю
    make_context_menu(text_widget)
    # Добавляем горячие клавиши для текстового поля
    add_hotkeys(text_widget)
    return text_widget


def clear_text(text_widget):
    """
    Функция для очистки текстового виджета (лога).

    :param text_widget: Виджет текста для очистки.
    """
    text_widget.config(state='normal')  # Разрешаем редактирование
    text_widget.delete(1.0, tk.END)  # Удаляем весь текст
    text_widget.config(state='disabled')  # Запрещаем редактирование


def create_plot(curves_info, X_label, Y_label, title, prY=False, savefile=False, file_plt='', fig=None, ax=None, legend=None):
    """Create a plot based on provided curve data.

    Parameters:
        curves_info (list[dict]): Список словарей с данными кривых. Каждый словарь должен
            содержать ключи ``'X_values'`` и ``'Y_values'``.
        X_label (str): Подпись оси X.
        Y_label (str): Подпись оси Y.
        title (str): Заголовок графика.
        prY (bool, optional): Если ``True``, значения Y отображаются в процентах.
        savefile (bool, optional): Флаг сохранения графика в файл.
        file_plt (str, optional): Путь к файлу для сохранения графика.
        fig (matplotlib.figure.Figure, optional): Существующая фигура для построения графика.
        ax (matplotlib.axes.Axes, optional): Существующая ось для построения графика.
        legend (bool, optional): Отображать легенду при наличии ``ax``.
    """
    if fig is None:
        if prY:
            fig = plt.figure(figsize=(8, 4.8))
            formatter = FuncFormatter(to_percent)
            plt.gca().yaxis.set_major_formatter(formatter)
        else:
            fig = plt.figure(figsize=(6.4, 4.8))
    if ax is None:
        for curve_info in curves_info:
            plt.plot(curve_info['X_values'], curve_info['Y_values'], marker=None, linestyle='-')
        plt.title(title, loc='left', fontsize=16, fontweight='bold')
        plt.xlabel(X_label)
        plt.ylabel(Y_label)
        plt.grid(True)
        fig.tight_layout()
        fig.subplots_adjust(bottom=0.15)
        if savefile:
            fig.savefig(file_plt)
        plt.close(fig)
    else:
        if legend:
            for curve_info in curves_info:
                ax.plot(curve_info['X_values'], curve_info['Y_values'], marker=None, linestyle='-', label=curve_info['curve_legend'])
        else:
            for curve_info in curves_info:
                ax.plot(curve_info['X_values'], curve_info['Y_values'], marker=None, linestyle='-')
        ax.set_title(title, fontsize=16, fontweight='bold', loc='left')
        # Обработка названий осей
        ax.set_xlabel(X_label)
        ax.set_ylabel(Y_label)
        ax.grid(True)
        fig.tight_layout()
        fig.subplots_adjust(bottom=0.15)
        # Условное добавление легенды в зависимости от состояния чекбокса
        if legend:
            ax.legend()


def create_txt_file_with_result(file_out_txt, data_for_file, log_text):
    with open(file_out_txt, 'w') as file:
        for name, graph in data_for_file.items():
            file.write(str(name) + '\n')
            for time_value in graph:
                file.write(str(time_value) + '\n')
            file.write('\n')
    message_log(log_text, "Создан текстовый файл со всеми параметрами")


def create_xlsx_file_with_result(file_out_xlsx, data_for_file, log_text):
    # Создаем новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active

    # Записываем данные в Excel
    row = 1
    for name, graph in data_for_file.items():
        names = name.split()
        for i, word in enumerate(names):
            header_cell = ws.cell(row=row, column=i + 1)
            header_cell.value = str(word)  # Записываем название
        row += 1
        for time_value in graph:
            for i, value in enumerate(time_value):
                if isinstance(value, str):
                    if value[-1] == '%':
                        value = value[:-1]
                ws.cell(row=row, column=i + 1).value = float(value)  # Записываем данные
            row += 1
        row += 1  # Пропускаем строку для разделения графов

    # Сохраняем файл
    wb.save(file_out_xlsx)
    message_log(log_text, "Создан эксель файл со всеми параметрами")


def create_png_plots(graph_with_time, file_path_outeig, log_text):
    for name, graph in graph_with_time.items():
        X_values = [float(item[0]) for item in graph]
        xlabel = 'Время $t$, с'
        namefig = FigureNames(name)

        # Список ключей для проверки в name
        keys = ['XR', 'YR', 'ZR', 'X', 'Y', 'Z']
        # Цикл для проверки каждого ключа в списке
        for key in keys:
            if key in name:
                directory = os.path.join(file_path_outeig, key)
                if not os.path.exists(directory):
                    os.makedirs(directory)
                file_plt = os.path.join(directory, f'{namefig.generate_filename()}.png')
                break  # Выход из цикла после первого совпадения
        if 'pr' in name:
            Y_values = [float(value[1].strip('%')) if isinstance(value[1], str) else value[1] for value in graph]
            curves_info = [{
                'X_values': X_values,
                'Y_values': Y_values
            }]
            create_plot(curves_info, xlabel, namefig.generate_plot_ylabel(), namefig.generate_plot_title(),
                        True, True, file_plt)
        else:
            Y_values = [float(item[1]) for item in graph]
            curves_info = [{
                'X_values': X_values,
                'Y_values': Y_values
            }]
            create_plot(curves_info, xlabel, namefig.generate_plot_ylabel(), namefig.generate_plot_title(),
                        False, True, file_plt)

    message_log(log_text, "Созданы графики всех характеристик")


def function1(log_text, path_entry):
    # Вывод текста в окно логов
    file_path_str = path_entry.get()
    if file_path_str:
        file_path = Path(file_path_str)
        message_log(log_text, f"Выбранный путь к проетку LS-Dyna: {file_path}")
        file_path_outeig = file_path / "eigresults"
        file_out_txt = file_path_outeig / "eigoutput.txt"
        file_out_xlsx = file_path_outeig / "eigoutput.xlsx"
        file_path_outeig.mkdir(parents=True, exist_ok=True)
        all_frequencies = []
        all_modal_mass = []
        time_data = []
        index = 1
        while True:
            file_name = f"eigout{index}"
            try:
                data_freq = extract_data_frec_from_file(file_path, file_name)
                frequencies = data_freq['cycles_values']
                modal_data = data_freq['modal_data']
                time = data_freq['time']
                time_data.append(time)
                all_modal_mass.append(modal_data)
                all_frequencies.append(frequencies)
                index += 1
            except FileNotFoundError:
                if index == 1:
                    message_log(log_text, "В данной папке нет расчетных файлов по частотному анализу")
                    return
                else:
                    n_file = index - 1
                    message_log(log_text, f"eigout{index} не найден. Значит последний eigout{n_file}")
                    message_log(log_text, "Точки во времени для модального анализа:")
                    message_log(log_text, str(time_data))
                break
        graphs = {
            "Xmass": [],
            "totalprXmass": [],
            "Ymass": [],
            "totalprYmass": [],
            "Zmass": [],
            "totalprZmass": [],
            "XRmass": [],
            "totalprXRmass": [],
            "YRmass": [],
            "totalprYRmass": [],
            "ZRmass": [],
            "totalprZRmass": [],
            "Xeig": [],
            "Yeig": [],
            "Zeig": [],
            "XReig": [],
            "YReig": [],
            "ZReig": [],
            "XN": [],
            "YN": [],
            "ZN": [],
            "XRN": [],
            "YRN": [],
            "ZRN": [],
            "prXmass": [],
            "prYmass": [],
            "prZmass": [],
            "prXRmass": [],
            "prYRmass": [],
            "prZRmass": [],
        }
        for mass, frequencies in zip(all_modal_mass, all_frequencies):
            onecoord = {
                "Xmass": 0,
                "Ymass": 0,
                "Zmass": 0,
                "XRmass": 0,
                "YRmass": 0,
                "ZRmass": 0,
            }
            for i, item in enumerate(mass):
                for key, value in item.items():
                    if key in ["X", "Y", "Z", "XR", "YR", "ZR"]:
                        if float(value) > onecoord[key + 'mass']:
                            onecoord[key + 'mass'] = float(value)
                            onecoord[key + 'eig'] = frequencies[int(item['N']) - 1]
                            onecoord[key + 'N'] = item['N']
                            onecoord['pr' + key + 'mass'] = item['pr' + key]
                            if i > 0:
                                onecoord['pr' + key + 'mass'] = str(float(onecoord['pr' + key + 'mass'].strip('%')) - float(
                                    mass[i - 1]['pr' + key].strip('%')))+'%'
                        onecoord['totalpr' + key + 'mass'] = item['pr' + key]

            for key, value in onecoord.items():
                if key in graphs:
                    graphs[key].append(value)
        data_for_file = {
            't XN Xeig Xmass prXmass totalprXmass': [],
            't YN Yeig Ymass prYmass totalprYmass': [],
            't ZN Zeig Zmass prZmass totalprZmass': [],
            't XRN XReig XRmass prXRmass totalprXRmass': [],
            't YRN YReig YRmass prYRmass totalprYRmass': [],
            't ZRN ZReig ZRmass prZRmass totalprZRmass': []
        }
        for i in range(n_file):
            for s in ['X', 'Y', 'Z', 'XR', 'YR', 'ZR']:
                data_for_file[f't {s}N {s}eig {s}mass pr{s}mass totalpr{s}mass'].append(
                    [time_data[i], graphs[f'{s}N'][i], graphs[f'{s}eig'][i], graphs[f'{s}mass'][i],
                     graphs[f'pr{s}mass'][i], graphs[f'totalpr{s}mass'][i]])

        graph_with_time = {
        }
        for key, graph in graphs.items():
            values_with_time = []
            for time, value in zip(time_data, graph):
                values_with_time.append([time, value])
            graph_with_time[key] = values_with_time
        # Создаем изображения графиков результатов
        create_png_plots(graph_with_time, file_path_outeig, log_text)

        # Создаем текстовый файл результатов
        create_txt_file_with_result(file_out_txt, data_for_file, log_text)
        # Создаем эксель файл результатов
        create_xlsx_file_with_result(file_out_xlsx, data_for_file, log_text)

    else:
        message_log(log_text, "Вы не выбрали путь к проетку LS-Dyna, выберите путь в ячейке выше...")


def on_combobox_event(event, *callbacks):
    """Вызывает все переданные функции-обработчики последовательно."""
    for callback in callbacks:
        if callable(callback):
            try:
                callback(event)
            except TypeError:
                callback()


def create_curve_box(input_frame, i, checkbox_var, saved_data):
    """Создает ячейку для настройки параметров кривой."""
    # Определяем высоту ячейки на основе состояния чекбокса
    dy = 190 if checkbox_var.get() else 130

    # Метка о параметрах кривой
    label_curve_box = ttk.Label(input_frame, text=f"Настройка параметров кривой {i}:")
    label_curve_box.place(x=10, y=0 + dy * (i - 1))

    # Метка о выборе типа кривой
    label_curve_type = ttk.Label(input_frame, text=f"Выберите тип кривой {i}:")
    label_curve_type.place(x=10, y=30 + dy * (i - 1))

    # Создание выпадающего меню для типа кривой
    combo_curve_type = ttk.Combobox(input_frame, values=["Частотный анализ", "2", "3"], state='readonly')
    combo_curve_type.place(x=250, y=30 + dy * (i - 1), width=150)
    combo_curve_type._name = f"curve_{i}_type"

    # Создание элементов для параметров X и Y
    label_curve_typeX = ttk.Label(input_frame, text="Выберите параметр для Х:")
    combo_curve_typeX = ttk.Combobox(input_frame, values=["Время", "Номер доминантной частота", "Частота",
                                                          "Масса", "Процент от общей массы", "Процент общей массы"],
                                     state='readonly')
    combo_curve_typeX._name = f"curve_{i}_typeXF"
    label_curve_typeY = ttk.Label(input_frame, text="Выберите параметр для Y:")
    combo_curve_typeY = ttk.Combobox(input_frame, values=["Время", "Номер доминантной частота", "Частота",
                                                          "Масса", "Процент от общей массы", "Процент общей массы"],
                                     state='readonly')
    combo_curve_typeY._name = f"curve_{i}_typeYF"
    label_curve_typeX_type = ttk.Label(input_frame, text="По какой оси:")
    combo_curve_typeX_type = ttk.Combobox(input_frame, values=["X", "Y", "Z", "XR", "YR", "ZR"],
                                          state='readonly')
    combo_curve_typeX_type._name = f"curve_{i}_typeXFtype"
    label_curve_typeY_type = ttk.Label(input_frame, text="По какой оси:")
    combo_curve_typeY_type = ttk.Combobox(input_frame, values=["X", "Y", "Z", "XR", "YR", "ZR"],
                                          state='readonly')
    combo_curve_typeY_type._name = f"curve_{i}_typeYFtype"

    # Установка позиций для параметров X и Y
    input_frame.update_idletasks()
    label_curve_typeX.place(x=combo_curve_type.winfo_x() + 170,
                            y=combo_curve_type.winfo_y() - 20)  # Отступ от типа кривой
    combo_curve_typeX.place(x=combo_curve_type.winfo_x() + 170,
                            y=combo_curve_type.winfo_y(), width=150)  # Позиция для X
    label_curve_typeX_type.place(x=combo_curve_type.winfo_x() + 170,
                                 y=combo_curve_type.winfo_y() + 25)  # Отступ от параметра X
    combo_curve_typeX_type.place(x=combo_curve_type.winfo_x() + 170,
                                 y=combo_curve_type.winfo_y() + 45, width=150)  # Позиция для оси X
    input_frame.update_idletasks()
    label_curve_typeY.place(x=combo_curve_typeX_type.winfo_x() + 170,
                            y=combo_curve_type.winfo_y() - 20)  # Отступ от параметра X
    combo_curve_typeY.place(x=combo_curve_typeX_type.winfo_x() + 170,
                            y=combo_curve_type.winfo_y(), width=150)  # Позиция для Y
    label_curve_typeY_type.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                 y=combo_curve_type.winfo_y() + 25)  # Отступ от параметра X
    combo_curve_typeY_type.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                 y=combo_curve_type.winfo_y() + 45, width=150)  # Позиция для оси Y

    # Привязка события изменения выбора в combo_curve_type
    combo_curve_type.bind("<<ComboboxSelected>>",
                          lambda event: on_combobox_event(event,
                                                          lambda e: on_combo_change_curve_type(input_frame,
                                                                                               combo_curve_type,
                                                                                               label_curve_typeX,
                                                                                               combo_curve_typeX,
                                                                                               label_curve_typeY,
                                                                                               combo_curve_typeY,
                                                                                               label_curve_typeX_type,
                                                                                               combo_curve_typeX_type,
                                                                                               label_curve_typeY_type,
                                                                                               combo_curve_typeY_type),
                                                          lambda e: saved_data[i - 1].update(
                                                              {'curve_type': combo_curve_type.get()})))

    # Метка для выбора файла с кривой
    label_path = ttk.Label(input_frame, text="Выберите файл с кривой:")
    label_path.place(x=10, y=90 + dy * (i - 1))

    # Создание текстового поля для ввода пути
    path_entry = create_text(input_frame, method="entry", height=1, state='normal', scrollbar=False)
    path_entry.place(x=10, y=110 + dy * (i - 1), width=600)

    path_entry._name = f"curve_{i}_filename"

    # Кнопка для выбора файла
    select_button = ttk.Button(input_frame, text="Выбор файла",
                               command=lambda: select_path(path_entry, path_type='file', saved_data=saved_data[i - 1]))
    select_button.place(x=620, y=108 + dy * (i - 1))

    # Проверка состояния чекбокса и соответствующая настройка интерфейса
    if checkbox_var.get():
        label_legend = ttk.Label(input_frame, text="Вбейте название кривой для легенды:")
        label_legend.place(x=10, y=150 + dy * (i - 1))

        legend_entry = create_text(input_frame, method="entry", height=1, state='normal', scrollbar=False)
        legend_entry.place(x=250, y=150 + dy * (i - 1), width=300)
        legend_entry._name = f"curve_{i}_legend"

        if i - 1 >= len(saved_data):
            saved_data.append({'curve_type': "", 'path': "", 'legend': "", 'curve_typeX': "", 'curve_typeY': "",
                               'curve_typeX_type': "", 'curve_typeY_type': ""})

        saved_curve_data = saved_data[i - 1]
        combo_curve_type.set(saved_curve_data['curve_type'])
        path_entry.insert(0, saved_curve_data['path'])
        legend_entry.insert(0, saved_curve_data['legend'])
        combo_curve_typeX.set(saved_curve_data['curve_typeX'])
        combo_curve_typeY.set(saved_curve_data['curve_typeY'])
        combo_curve_typeX_type.set(saved_curve_data['curve_typeX_type'])
        combo_curve_typeY_type.set(saved_curve_data['curve_typeY_type'])

        # Сохранение изменений
        combo_curve_typeX.bind("<<ComboboxSelected>>",
                               lambda e: saved_data[i - 1].update({'curve_typeX': combo_curve_typeX.get()}))
        combo_curve_typeY.bind("<<ComboboxSelected>>",
                               lambda e: saved_data[i - 1].update({'curve_typeY': combo_curve_typeY.get()}))
        combo_curve_typeX_type.bind("<<ComboboxSelected>>", lambda e: saved_data[i - 1].update(
            {'curve_typeX_type': combo_curve_typeX_type.get()}))
        combo_curve_typeY_type.bind("<<ComboboxSelected>>", lambda e: saved_data[i - 1].update(
            {'curve_typeY_type': combo_curve_typeY_type.get()}))
        path_entry.bind("<FocusOut>", lambda e: saved_data[i - 1].update({'path': path_entry.get()}))
        legend_entry.bind("<FocusOut>", lambda e: saved_data[i - 1].update({'legend': legend_entry.get()}))
    else:
        if i - 1 >= len(saved_data):
            saved_data.append({'curve_type': "", 'path': "", 'legend': "", 'curve_typeX': "", 'curve_typeY': "",
                               'curve_typeX_type': "", 'curve_typeY_type': ""})

        saved_curve_data = saved_data[i - 1]
        combo_curve_type.set(saved_curve_data['curve_type'])
        path_entry.insert(0, saved_curve_data['path'])
        combo_curve_typeX.set(saved_curve_data['curve_typeX'])
        combo_curve_typeY.set(saved_curve_data['curve_typeY'])
        combo_curve_typeX_type.set(saved_curve_data['curve_typeX_type'])
        combo_curve_typeY_type.set(saved_curve_data['curve_typeY_type'])

        # Сохранение изменений
        combo_curve_typeX.bind("<<ComboboxSelected>>",
                               lambda e: saved_data[i - 1].update({'curve_typeX': combo_curve_typeX.get()}))
        combo_curve_typeY.bind("<<ComboboxSelected>>",
                               lambda e: saved_data[i - 1].update({'curve_typeY': combo_curve_typeY.get()}))
        combo_curve_typeX_type.bind("<<ComboboxSelected>>", lambda e: saved_data[i - 1].update(
            {'curve_typeX_type': combo_curve_typeX_type.get()}))
        combo_curve_typeY_type.bind("<<ComboboxSelected>>", lambda e: saved_data[i - 1].update(
            {'curve_typeY_type': combo_curve_typeY_type.get()}))
        path_entry.bind("<FocusOut>", lambda e: saved_data[i - 1].update({'path': path_entry.get()}))

    on_combo_change_curve_type(input_frame, combo_curve_type, label_curve_typeX, combo_curve_typeX, label_curve_typeY,
                               combo_curve_typeY, label_curve_typeX_type, combo_curve_typeX_type,
                               label_curve_typeY_type,
                               combo_curve_typeY_type)


def update_curves(frame, num_curves, next_frame, checkbox_var, saved_data):
    """Обновляет кривые в соответствии с выбранным количеством и состоянием чекбокса."""
    # Очищаем старые виджеты
    for widget in frame.winfo_children():
        widget.destroy()

    if num_curves == '':
        return
    else:
        num_curves_int = int(num_curves)

    # Меняем высоту фрейма в зависимости от количества кривых
    frame_height = 210 * num_curves_int if checkbox_var.get() else 150 * num_curves_int
    frame.place_configure(height=frame_height)

    # Восстанавливаем данные, если они есть
    for i in range(len(saved_data), num_curves_int):
        saved_data.append({'curve_type': "", 'path': "", 'legend': "", 'curve_typeX': "", 'curve_typeY': "",
                           'curve_typeX_type': "", 'curve_typeY_type': ""})

    for i in range(1, num_curves_int + 1):
        create_curve_box(frame, i, checkbox_var, saved_data)

    next_frame.place(x=10, y=frame.winfo_y() + frame_height + 10)  # Обновляем координаты следующего фрейма


def save_file(entry_widget, graph_info):
    file_name = entry_widget.get()
    if file_name:
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                                                 initialfile=file_name)
        if file_path:
            try:
                curves_info = [{
                    'X_values': graph_info['X_values'],
                    'Y_values': graph_info['Y_values']
                }]
                create_plot(curves_info, graph_info['X_label'], graph_info['Y_label'], graph_info['title'],
                            savefile=True,
                            file_plt=file_path)  # Генерация и сохранение графика
                messagebox.showinfo("Успех", f"График сохранен: {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")
    else:
        messagebox.showerror("Ошибка", "Имя файла не может быть пустым!")


def on_combo_change_curve_type(frame, combo, label_curve_typeX, combo_curve_typeX, label_curve_typeY, combo_curve_typeY,
                               label_curve_typeX_type, combo_curve_typeX_type, label_curve_typeY_type,
                               combo_curve_typeY_type):
    if combo.get() == "Частотный анализ":
        if not label_curve_typeX.winfo_viewable():  # Проверяем, отображается ли виджет
            frame.update_idletasks()
            label_curve_typeX.place(x=combo.winfo_x() + 170,
                                    y=combo.winfo_y() - 20)  # Отступ от типа кривой

            combo_curve_typeX.place(x=combo.winfo_x() + 170,
                                    y=combo.winfo_y(), width=150)  # Позиция для X
            label_curve_typeX_type.place(x=combo.winfo_x() + 170,
                                         y=combo.winfo_y() + 25)  # Отступ от параметра X
            combo_curve_typeX_type.place(x=combo.winfo_x() + 170,
                                         y=combo.winfo_y() + 45, width=150)  # Позиция для Y
            frame.update_idletasks()
            label_curve_typeY.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                    y=combo.winfo_y() - 20)  # Отступ от параметра X
            combo_curve_typeY.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                    y=combo.winfo_y(), width=150)  # Позиция для Y
            label_curve_typeY_type.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                         y=combo.winfo_y() + 25)  # Отступ от параметра X
            combo_curve_typeY_type.place(x=combo_curve_typeX_type.winfo_x() + 170,
                                         y=combo.winfo_y() + 45, width=150)  # Позиция для Y
    else:
        label_curve_typeX.place_forget()
        combo_curve_typeX.place_forget()
        label_curve_typeY.place_forget()
        combo_curve_typeY.place_forget()
        label_curve_typeX_type.place_forget()
        combo_curve_typeX_type.place_forget()
        label_curve_typeY_type.place_forget()
        combo_curve_typeY_type.place_forget()





class AxisTitleProcessor:
    def __init__(self, combo_title, combo_size):
        self.combo_title = combo_title
        self.combo_size = combo_size
        self.title_mapping = {
            "Время": "Время $t$",
            "Частота 1": "Частота ${{f}}_{{\mathit{{1}}}}$",
            # Добавьте другие варианты здесь при необходимости
        }

    def get_processed_title(self):
        title = self.combo_title.get()
        size = self.combo_size.get()
        processed_title = self.title_mapping.get(title, title)
        if title in self.title_mapping:
            return f"{processed_title}, {size}"
        return processed_title


def get_X_Y_data(curve_info):
    if curve_info['curve_type'] == 'Частотный анализ':
        try:
            # Открытие файла
            with open(curve_info['curve_file'], 'r') as file:
                lines = file.readlines()

            # Динамическое определение заголовков на основе curve_typeXF_type и curve_typeYF_type
            header_XF = f"t {curve_info['curve_typeXF_type']}N {curve_info['curve_typeXF_type']}eig {curve_info['curve_typeXF_type']}mass pr{curve_info['curve_typeXF_type']}mass totalpr{curve_info['curve_typeXF_type']}mass"
            header_YF = f"t {curve_info['curve_typeYF_type']}N {curve_info['curve_typeYF_type']}eig {curve_info['curve_typeYF_type']}mass pr{curve_info['curve_typeYF_type']}mass totalpr{curve_info['curve_typeYF_type']}mass"

            # Динамическое определение индекса столбца для каждого параметра curve_typeXF и curve_typeYF
            headers_map = {
                "Время": 0,
                "Номер доминантной частота": 1,
                "Частота": 2,
                "Масса": 3,
                "Процент от общей массы": 4,
                "Процент общей массы": 5
            }

            # Определяем, какой столбец нужно извлекать для X и Y в зависимости от curve_typeXF и curve_typeYF
            index_X = headers_map.get(curve_info['curve_typeXF'])
            index_Y = headers_map.get(curve_info['curve_typeYF'])

            if index_X is None or index_Y is None:
                logger.error("Ошибка: некорректные параметры curve_typeXF или curve_typeYF.")
                return

            # Переменные для хранения данных
            X_data = []
            Y_data = []
            current_block_X = False
            current_block_Y = False

            for line in lines:
                line = line.strip()

                # Определение начала блока
                if line == header_XF and line == header_YF:
                    current_block_X = True
                    current_block_Y = True
                    continue
                elif line == header_XF:
                    current_block_X = True
                    current_block_Y = False
                    continue
                elif line == header_YF:
                    current_block_Y = True
                    current_block_X = False
                    continue
                elif line == '':
                    current_block_X = False
                    current_block_Y = False
                    continue

                # Извлечение данных из блока X
                if current_block_X:
                    try:
                        data_list = ast.literal_eval(line)  # Преобразуем строку в список
                        if isinstance(data_list, list) and len(data_list) > index_X:
                            if index_X == 4 or index_X == 5:
                                X_data.append(float(data_list[index_X].strip('%')))
                            else:
                                X_data.append(float(data_list[index_X]))  # Извлекаем данные для X
                    except (ValueError, SyntaxError):
                        # Если строку нельзя преобразовать в список, пропускаем её
                        logger.error("Ошибка преобразования строки: %s", line)

                # Извлечение данных из блока Y
                if current_block_Y:
                    try:
                        data_list = ast.literal_eval(line)  # Преобразуем строку в список
                        if isinstance(data_list, list) and len(data_list) > index_Y:
                            if index_Y == 4 or index_Y == 5:
                                logger.debug("%s", data_list[index_Y])
                                Y_data.append(float(data_list[index_Y].strip('%')))
                            else:
                                Y_data.append(float(data_list[index_Y]))  # Извлекаем данные для Y
                    except (ValueError, SyntaxError):
                        # Если строку нельзя преобразовать в список, пропускаем её
                        logger.error("Ошибка преобразования строки: %s", line)

            # Сохранение данных в curve_info
            curve_info['X_values'] = X_data
            curve_info['Y_values'] = Y_data

        except FileNotFoundError:
            logger.error("Файл '%s' не найден.", curve_info['curve_file'])
        except IOError:
            logger.error("Ошибка при чтении файла '%s'.", curve_info['curve_file'])


def generate_graph(ax,fig, canvas, path_entry_title, combo_titleX, combo_titleX_size, combo_titleY, combo_titleY_size,
                   legend_checkbox, curves_frame, combo_curves):
    # Очистка предыдущего графика
    ax.clear()
    # Считываем заголовок из поля ввода
    title = path_entry_title.get()
    xlabel_processor = AxisTitleProcessor(combo_titleX, combo_titleX_size)
    ylabel_processor = AxisTitleProcessor(combo_titleY, combo_titleY_size)
    xlabel = xlabel_processor.get_processed_title()
    ylabel = ylabel_processor.get_processed_title()

    # Считываем количество кривых из combobox
    num_curves = int(combo_curves.get())

    # Генерация данных для графика (пример - синусоида)
    x = np.linspace(0, 10, 100)
    curves_info = []
    # Построение каждой кривой в цикле
    for i in range(1, num_curves + 1):
        curve_info = {}
        for widget in curves_frame.winfo_children():
            if hasattr(widget, '_name'):
                widget_name = widget._name

                # Проверяем тип кривой
                if widget_name == f"curve_{i}_type":
                    curve_info['curve_type'] = widget.get()

                # Если тип кривой "Частотный анализ", собираем дополнительные данные
                if 'curve_type' in curve_info and curve_info['curve_type'] == "Частотный анализ":
                    if widget_name == f"curve_{i}_typeXF":
                        curve_info['curve_typeXF'] = widget.get()
                    elif widget_name == f"curve_{i}_typeYF":
                        curve_info['curve_typeYF'] = widget.get()
                    elif widget_name == f"curve_{i}_typeXFtype":
                        curve_info['curve_typeXF_type'] = widget.get()
                    elif widget_name == f"curve_{i}_typeYFtype":
                        curve_info['curve_typeYF_type'] = widget.get()

                # Получаем имя файла для каждой кривой
                if widget_name == f"curve_{i}_filename":
                    curve_info['curve_file'] = widget.get()

                # Проверяем наличие легенды, если отмечен чекбокс
                if legend_checkbox.get() and widget_name == f"curve_{i}_legend":
                    curve_info['curve_legend'] = widget.get()

        # Добавляем информацию о кривой в общий список
        get_X_Y_data(curve_info)
        curves_info.append(curve_info)

    create_plot(curves_info, xlabel, ylabel, title,
                fig=fig, ax=ax,legend=legend_checkbox.get())

    # Перерисовка графика
    canvas.draw()


# Функция завершения программы при закрытии окна
def on_closing(root):
    root.quit()  # Останавливаем цикл Tkinter
    root.destroy()  # Уничтожаем окно


def main():
    configure_matplotlib()
    # Создание главного окна
    root = tk.Tk()
    root.geometry("1500x1100")
    root.title("Работа с графиками")
    root.resizable(True, True)

    # Создание вкладок
    notebook = ttk.Notebook(root)
    notebook.pack(expand=True, fill=tk.BOTH)

    # Добавляем вкладки
    create_tab1(notebook, create_text, update_curves, generate_graph, save_file)
    create_tab2(notebook)

    create_tab3(notebook, create_text, function1, clear_text)

    # Привязка обработчика закрытия окна
    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root))

    # Запуск окна
    root.mainloop()


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    main()
