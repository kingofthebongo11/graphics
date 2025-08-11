import csv
import logging
from pathlib import Path
from tkinter import messagebox

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

logger = logging.getLogger(__name__)


def read_X_Y_from_excel(curve_info):
    try:
        path = Path(curve_info['curve_file'])
        suffix = path.suffix.lower()
        X_data = []
        Y_data = []
        horizontal = curve_info.get('horizontal', False)
        use_offset = curve_info.get('use_offset', False)
        h_off = int(curve_info.get('offset_horizontal', 0)) if use_offset else 0
        v_off = int(curve_info.get('offset_vertical', 0)) if use_offset else 0
        use_ranges = curve_info.get('use_ranges', False)
        range_x = curve_info.get('range_x', '')
        range_y = curve_info.get('range_y', '')

        def split_sheet_range(rng: str):
            if '!' in rng:
                sheet, cells = rng.split('!', 1)
                return sheet, cells
            return None, rng

        if use_ranges and (range_x or range_y):
            if suffix in {'.xlsx', '.xlsm'}:
                wb = load_workbook(path, read_only=True, data_only=True)

                def read_cells(rng: str):
                    sheet, cells = split_sheet_range(rng)
                    ws = wb[sheet] if sheet else wb.active
                    try:
                        return [cell.value for row in ws[cells] for cell in row]
                    except ValueError:
                        messagebox.showerror("Ошибка", "Неверный формат диапазона Excel")
                        raise

                try:
                    if range_x and range_y:
                        values_x = read_cells(range_x)
                        values_y = read_cells(range_y)
                        for x, y in zip(values_x, values_y):
                            if x is None or y is None:
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                            try:
                                X_data.append(float(str(x).replace(',', '.')))
                                Y_data.append(float(str(y).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                    elif range_x:
                        for x in read_cells(range_x):
                            if x is None:
                                logger.error("Некорректное значение данных: %s", x)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x}")
                                return
                            try:
                                X_data.append(float(str(x).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректное значение данных: %s", x)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x}")
                                return
                    else:
                        for y in read_cells(range_y):
                            if y is None:
                                logger.error("Некорректное значение данных: %s", y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {y}")
                                return
                            try:
                                Y_data.append(float(str(y).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректное значение данных: %s", y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {y}")
                                return
                except ValueError:
                    return

            elif suffix == '.csv':
                with open(path, 'r', encoding='utf-8') as f:
                    rows = list(csv.reader(f))

                def get_range_vals(rng):
                    _, cells = split_sheet_range(rng)
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(cells)
                    except ValueError:
                        messagebox.showerror("Ошибка", "Неверный формат диапазона Excel")
                        raise
                    vals = []
                    for r_idx in range(min_row, max_row + 1):
                        if r_idx - 1 >= len(rows):
                            break
                        row = rows[r_idx - 1]
                        for c_idx in range(min_col, max_col + 1):
                            if c_idx - 1 < len(row):
                                vals.append(row[c_idx - 1])
                    return vals

                try:
                    if range_x and range_y:
                        values_x = get_range_vals(range_x)
                        values_y = get_range_vals(range_y)
                        for x, y in zip(values_x, values_y):
                            if x is None or y is None or x == '' or y == '':
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                            try:
                                X_data.append(float(str(x).replace(',', '.')))
                                Y_data.append(float(str(y).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                    elif range_x:
                        for x in get_range_vals(range_x):
                            if x is None or x == '':
                                logger.error("Некорректное значение данных: %s", x)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x}")
                                return
                            try:
                                X_data.append(float(str(x).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректное значение данных: %s", x)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x}")
                                return
                    else:
                        for y in get_range_vals(range_y):
                            if y is None or y == '':
                                logger.error("Некорректное значение данных: %s", y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {y}")
                                return
                            try:
                                Y_data.append(float(str(y).replace(',', '.')))
                            except (ValueError, TypeError):
                                logger.error("Некорректное значение данных: %s", y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {y}")
                                return
                except ValueError:
                    return
            else:
                logger.error("Неподдерживаемый формат файла: %s", suffix)
                return

            curve_info['X_values'] = X_data
            curve_info['Y_values'] = Y_data
            return

        if suffix in {'.xlsx', '.xlsm'}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            if horizontal:
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= v_off + 2:
                    row_x, row_y = rows[v_off], rows[v_off + 1]
                    for x, y in zip(row_x[h_off:], row_y[h_off:]):
                        if x is None or y is None:
                            logger.error("Некорректная пара данных: %s, %s", x, y)
                            messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                            return
                        try:
                            X_data.append(float(str(x).replace(',', '.')))
                            Y_data.append(float(str(y).replace(',', '.')))
                        except (ValueError, TypeError):
                            logger.error("Некорректная пара данных: %s, %s", x, y)
                            messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                            return
            else:
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx < v_off:
                        continue
                    if row is None or len(row) <= h_off + 1:
                        logger.error("Некорректная строка данных: %s", row)
                        messagebox.showerror("Ошибка", f"Некорректные данные в строке: {row}")
                        return
                    x, y = row[h_off], row[h_off + 1]
                    if x is None or y is None:
                        logger.error("Некорректная строка данных: %s", row)
                        messagebox.showerror("Ошибка", f"Некорректные данные в строке: {row}")
                        return
                    try:
                        X_data.append(float(str(x).replace(',', '.')))
                        Y_data.append(float(str(y).replace(',', '.')))
                    except (ValueError, TypeError):
                        logger.error("Некорректная строка данных: %s", row)
                        messagebox.showerror("Ошибка", f"Некорректные данные в строке: {row}")
                        return
        elif suffix == '.csv':
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                if horizontal:
                    rows = list(reader)
                    if len(rows) >= v_off + 2:
                        row_x = rows[v_off]
                        row_y = rows[v_off + 1]
                        for x, y in zip(row_x[h_off:], row_y[h_off:]):
                            if not x or not y:
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                            try:
                                X_data.append(float(x.replace(',', '.')))
                                Y_data.append(float(y.replace(',', '.')))
                            except ValueError:
                                logger.error("Некорректная пара данных: %s, %s", x, y)
                                messagebox.showerror("Ошибка", f"Некорректные данные в строке: {x} {y}")
                                return
                else:
                    for idx, row in enumerate(reader):
                        if idx < v_off:
                            continue
                        if len(row) <= h_off + 1:
                            logger.error("Некорректная строка данных: %s", row)
                            messagebox.showerror("Ошибка", f"Некорректные данные в строке: {' '.join(row)}")
                            return
                        x = row[h_off]
                        y = row[h_off + 1]
                        if not x or not y:
                            logger.error("Некорректная строка данных: %s", row)
                            messagebox.showerror("Ошибка", f"Некорректные данные в строке: {' '.join(row)}")
                            return
                        try:
                            X_data.append(float(x.replace(',', '.')))
                            Y_data.append(float(y.replace(',', '.')))
                        except ValueError:
                            logger.error("Некорректная строка данных: %s", row)
                            messagebox.showerror("Ошибка", f"Некорректные данные в строке: {' '.join(row)}")
                            return
        else:
            logger.error("Неподдерживаемый формат файла: %s", suffix)
            return

        curve_info['X_values'] = X_data
        curve_info['Y_values'] = Y_data
    except FileNotFoundError:
        logger.error("Файл '%s' не найден.", curve_info['curve_file'])
        messagebox.showerror("Ошибка", f"Не удалось открыть файл {path}")
    except Exception:
        logger.error("Ошибка при чтении файла '%s'.", curve_info['curve_file'])
        messagebox.showerror("Ошибка", f"Не удалось открыть файл {path}")
