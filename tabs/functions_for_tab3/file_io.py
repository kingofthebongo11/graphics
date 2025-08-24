import openpyxl
from pathlib import Path
from widgets.message_log import message_log


def create_txt_file_with_result(file_out_txt, data_for_file, log_text):
    """Создает текстовый файл с результатами частотного анализа."""
    file_out_txt = Path(file_out_txt).resolve()
    with file_out_txt.open('w') as file:
        for name, graph in data_for_file.items():
            file.write(str(name) + '\n')
            for time_value in graph:
                file.write(str(time_value) + '\n')
            file.write('\n')
    message_log(log_text, "Создан текстовый файл со всеми параметрами")


def create_xlsx_file_with_result(file_out_xlsx, data_for_file, log_text):
    """Создает Excel-файл с результатами частотного анализа."""
    file_out_xlsx = Path(file_out_xlsx).resolve()
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    for name, graph in data_for_file.items():
        names = name.split()
        for i, word in enumerate(names):
            header_cell = ws.cell(row=row, column=i + 1)
            header_cell.value = str(word)
        row += 1
        for time_value in graph:
            for i, value in enumerate(time_value):
                if isinstance(value, str) and value.endswith('%'):
                    value = value[:-1]
                ws.cell(row=row, column=i + 1).value = float(value)
            row += 1
        row += 1

    wb.save(file_out_xlsx)
    message_log(log_text, "Создан эксель файл со всеми параметрами")
