import tempfile
from openpyxl import Workbook

from tabs.functions_for_tab1.curves_from_file.combined_curve import read_X_Y_from_combined


def test_read_combined_excel_ranges():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 1
    ws['A2'] = 2
    ws['A3'] = 3
    ws['B1'] = 4
    ws['B2'] = 5
    ws['B3'] = 6
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        curve_info = {
            'X_source': {
                'source': 'Excel файл',
                'curve_file': tmp.name,
                'range_x': 'A1:A3',
                'use_ranges': True,
                'column': 0,
            },
            'Y_source': {
                'source': 'Excel файл',
                'curve_file': tmp.name,
                'range_y': 'B1:B3',
                'use_ranges': True,
                'column': 1,
            },
        }
        read_X_Y_from_combined(curve_info)
        assert curve_info['X_values'] == [1.0, 2.0, 3.0]
        assert curve_info['Y_values'] == [4.0, 5.0, 6.0]
