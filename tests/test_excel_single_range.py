import tempfile
import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))
from tabs.functions_for_tab1.curves_from_file.excel_file import read_X_Y_from_excel


def test_read_single_x_range():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 1
    ws['A2'] = 2
    ws['A3'] = 3
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        info = {
            'curve_file': tmp.name,
            'use_ranges': True,
            'range_x': 'A1:A3',
            'range_y': '',
        }
        read_X_Y_from_excel(info)
        assert info['X_values'] == [1.0, 2.0, 3.0]
        assert info['Y_values'] == []
